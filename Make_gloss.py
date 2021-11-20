import openpyxl
from openpyxl.cell.cell import VALID_TYPES
from botok.modifytokens.splitaffixed import split_affixed
from pybo.utils.regex_batch_apply import batch_apply_regex, get_regex_pairs
from botok.tokenizers.wordtokenizer import WordTokenizer
import csv
from pathlib import Path
import re
import shutil
from tempfile import NamedTemporaryFile


def extract_valid_column(valid_column, wb_obj, sheet_obj, f_writer, rules):

    row = sheet_obj.max_row
    column = sheet_obj.max_column

    wt = WordTokenizer()

    for i in range(2, row+1):
        gloss_list = []
        for j in valid_column:
            cell_obj = sheet_obj.cell(row=i, column=j)
            gloss_list.append(cell_obj.value)
        to_transifix_csv_format(gloss_list, f_writer, wt, rules)


def to_transifix_csv_format(gloss_list, f_writer, wt, rules):

    transifex_list = []
    pattern = "་$"

    # adding tibetan to term
    if len(gloss_list[7]) != 0:
        tibetan_tokenized, pos = tokenize_line(gloss_list[7], wt, rules)
        tibetan_tokenized = tibetan_tokenized.replace(" །", "")
        tibetan_tokenized = re.sub(pattern, "", tibetan_tokenized)
    else:
        tibetan_tokenized = gloss_list[7]

    transifex_list.append(tibetan_tokenized)

    # Adding demo pos to POS
    pos = ""
    transifex_list.append(pos)

    # Adding Type-Sanskirt to comment
    comment = str(gloss_list[0])+"-"+str(gloss_list[8])
    transifex_list.append(comment)

    # adding yes to is_case_sensitive
    transifex_list.append("yes")

    # Adding english to en
    transifex_list.append(gloss_list[5])

    # Adding defintion to comment_en
    comment = r"\n"+str(gloss_list[4])
    comment += r"\n"+str(gloss_list[1])+","+str(gloss_list[2]).replace(
        "Translated by the", "").replace("Translated by", "")+","+str(gloss_list[3])

    transifex_list.append(comment)

    parse_same_term(transifex_list)


def parse_same_term(transifex_list):
    filename = "sample_glossary.csv"
    temp_file = NamedTemporaryFile(delete=False, mode='w')
    set = 0
    header = ["term", "pos", "comment",
              "is_case_sensitive", "en", "comment_en"]

    with open(filename, "r") as csvfile, temp_file:
        rows = []
        reader = csv.DictReader(csvfile)
        writer = csv.DictWriter(temp_file, fieldnames=header)
        writer.writeheader()
        if len(list(reader)) != 0:
            csvfile.seek(0)
            for row in reader:
                if row["term"] == "term":
                    continue
                if row["term"] == transifex_list[0] and row["en"] == transifex_list[4]:
                    writer.writerow({
                        "term": row["term"],
                        "pos": row["pos"],
                        "comment": row["comment"],
                        "is_case_sensitive": row["is_case_sensitive"],
                        "en": row["en"],
                        "comment_en": row["comment_en"] + transifex_list[5]
                    })
                    set = 1
                elif row["term"] == transifex_list[0] and row["en"] != transifex_list[4]:
                    writer.writerow({
                        "term": row["term"],
                        "pos": row["pos"],
                        "comment": row["comment"],
                        "is_case_sensitive": row["is_case_sensitive"],
                        "en": row["en"],
                        "comment_en": row["comment_en"] + transifex_list[4]+transifex_list[5]
                    })
                    set = 1
                else:
                    writer.writerow({
                        "term": row["term"],
                        "pos": row["pos"],
                        "comment": row["comment"],
                        "is_case_sensitive": row["is_case_sensitive"],
                        "en": row["en"],
                        "comment_en": row["comment_en"]
                    })
                    
        if set != 1:
            writer.writerow({
                "term": transifex_list[0],
                "pos": transifex_list[1],
                "comment": transifex_list[2],
                "is_case_sensitive": transifex_list[3],
                "en": transifex_list[4],
                "comment_en": transifex_list[5]
            })
    shutil.move(temp_file.name, filename)


def tokenize_line(line, wt, rules):

    new_line = ''
    pos = ""
    tokens = wt.tokenize(line, split_affixes=True)

    for token in tokens:
        new_line += f'{token.text} '
    new_line = new_line.strip()
    normalized_line = normalize_line(new_line, rules)

    if len(tokens) == 1:
        for token in tokens:
            pos = token.pos

    return normalized_line, pos


def normalize_line(line, rules):
    normalized_line = batch_apply_regex(line, rules)
    return normalized_line


def write_csv(path, valid_column, rules):
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    header = ["term", "pos", "comment",
              "is_case_sensitive", "en", "comment_en"]
    with open("sample_glossary.csv", "w") as f:
        f_writer = csv.DictWriter(f, fieldnames=header)
        f_writer.writeheader()
        extract_valid_column(valid_column, wb_obj, sheet_obj, f_writer, rules)


if __name__ == "__main__":
    path = "84000_Glossary_09-24-2021-full.xlsx"
    regex_file = Path('./regex.txt')
    rules = get_regex_pairs(regex_file.open(encoding="utf-8-sig").readlines())
    valid_column = (1, 2, 4, 5, 6, 7, 8, 9, 12)
    write_csv(path, valid_column, rules)
